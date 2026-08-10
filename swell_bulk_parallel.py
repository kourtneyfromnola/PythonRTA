from __future__ import annotations

"""
SWELL bulk parallel runner.

Input workbook sheets:
  - Matrix
  - ShrinkYield and Timing
  - Gross Cost
  - Type Curves
Optional:
  - Defaults

This script:
  1) reads all independent Matrix lists,
  2) filters LL-Zone-WPS-GSA combinations to only those with a type curve,
  3) crosses each valid type curve with every operator and county,
  4) resolves operator assumptions and gross costs,
  5) writes worker-specific WellAttributes + Type Curves into separate SWELL copies,
  6) runs independent Excel instances in parallel,
  7) outputs Case ID, Operator, Lateral Length, WPS, Zone, GSA, County, NPV, AARR, CoS.

No basin names are hard-coded. Create analogous input workbooks for Midland,
Delaware TX, etc. and use the same code.

Requested fixed assumptions:
  WI = 100%
  Capital WI = 100%
  NRI = 75%
  MMBTU/MCF = 1
  Gross Investment = drilling + facilities + flowlines + completion
  Artificial Lift Gross = 0
  Spud Date = 1/1/2027
  Artificial Lift Years After POP = 1
  Obligation = REQUIRED WELL
  TC Adjustment = 1
  WOR1 = WOR2 = type-curve WOR; WOR1 length = 1 month

Lateral code mapping:
  100 -> 5000 ft
  150 -> 7500 ft
  200 -> 10000 ft
  250 -> 12500 ft
  300 -> 13000 ft

IMPORTANT: validate the 5% test against known manual SWELL cases first.
"""

import argparse
import itertools
import math
import os
import random
import re
import shutil
import tempfile
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

LATERAL_FEET_MAP = {100: 5000, 150: 7500, 200: 10000, 250: 12500, 300: 13000}
XL_CALCULATION_MANUAL = -4135
XL_DONE = 0

# ---------------------------------------------------------------------------
# SWELL WORKBOOK CONFIG
# These are the only things you should ever need to touch if your SWELL
# model uses different Named Range / macro / sheet names than assumed below.
# Run:  python swell_bulk_parallel.py --diagnose --swell "path\to\SWELL.xlsm"
# to print the actual Named Ranges, macros, and ListObjects found in your
# SWELL workbook so you can correct any mismatches here BEFORE a full run.
# ---------------------------------------------------------------------------
SWELL_CONFIG = {
    "case_name": "case",
    "npv_name": "npv",
    "aarr_name": "aarr",
    "cos_input_name": "flatOilInput",
    "price_scenario_name": "priceScenario",
    "econ_limit_name": "econLimit",
    "price_scenario_value": "LRP P50",
    "econ_limit_value": "YES",
    "cos_input_value": 40,
    "macro_candidates": ["CoS_Calc", "CoS.CoS_Calc"],
    "dashboard_sheet": "SW Dashboard",
}

HEADER_ALIASES = {
    "operators": ["operators", "operator"],
    "lateral_lengths": ["lateral lengths", "lateral length", "ll"],
    "wps": ["wps"],
    "zone": ["zone", "formation", "reservoir"],
    "gsa": ["gsa"],
    "county": ["county", "counties"],
    "yield": ["yield bbl mmcf", "yield bbl/mmcf", "ngl yield"],
    "shrink": ["gas shrink fraction", "gas shrink", "shrink"],
    "drill_month": ["drilling months after spud", "drilling month after spud"],
    "fac_month": ["facilities months after spud", "facility months after spud"],
    "flow_month": ["flowlines months after spud", "flowline months after spud"],
    "comp_month": ["completion months after spud", "complet months after spud"],
    "pop_month": ["pop/opex months after spud", "pop opex months after spud", "pop months after spud"],
    "drill_cost": ["drilling gross m", "drilling gross m$", "drilling $m gross"],
    "fac_cost": ["facilities gross m", "facilities gross m$", "facilities $m gross"],
    "flow_cost": ["flowlines gross m", "flowlines gross m$", "flowlines $m gross"],
    "comp_cost": ["completion gross m", "completion gross m$", "complet $m gross"],
    "formation": ["formation", "zone"],
    "tc_ll": ["ll", "lateral length", "lateral lengths"],
    "wor": ["wor"],
    "tc_address": ["type curve address", "typecurve address", "type curve", "typecurve"],
}

WELLATTR_ALIASES = {
    "case": ["case"],
    "api": ["api"],
    "project_name": ["well/project name", "wellproject name", "well project name"],
    "well_count": ["well count", "wellcount"],
    "type_curve": ["type curve", "typecurve"],
    "reservoir": ["reservoir"],
    "operator": ["operator"],
    "wi": ["wi"],
    "capital_wi": ["capital wi", "capitalwi"],
    "nri": ["nri"],
    "yield": ["ngl yield bbl/mmcf", "ngl yield", "yield bbl/mmcf"],
    "shrink": ["gas shrink", "gas shrink fraction"],
    "mmbtu": ["mmbtu/mcf", "mmbtu mcf"],
    "lateral_ft": ["lateral length ft", "lateral length"],
    "drill_cost": ["drilling $m gross", "drilling gross m$", "drilling gross"],
    "fac_cost": ["facilities $m gross", "facilities gross m$", "facilities gross"],
    "flow_cost": ["flowlines $m gross", "flowlines gross m$", "flowlines gross"],
    "comp_cost": ["complet $m gross", "completion $m gross", "completion gross m$", "complet gross"],
    "art_lift_cost": ["artificial lift $m gross", "art lift $m gross", "artificial lift gross"],
    "gross_invest": ["gross invest", "gross investment"],
    "spud_date": ["spud date", "spuddate"],
    "drill_month": ["drilling months after spud", "drilling month after spud"],
    "fac_month": ["facilities months after spud", "facility months after spud"],
    "flow_month": ["flowlines months after spud", "flowline months after spud"],
    "comp_month": ["complet months after spud", "completion months after spud"],
    "pop_month": ["pop/opex months after spud", "pop opex months after spud"],
    "art_lift_years": ["art. lift years after pop", "art lift years after pop", "artificial lift years after pop"],
    "obligation": ["obligation"],
    "tc_adj": ["tc adj", "tc adjustment", "tc adjust"],
    "county": ["county"],
    "wor1": ["wor 1", "wor1"],
    "wor1_length": ["wor 1 length (mo.)", "wor 1 length mo", "wor1 length"],
    "wor2": ["wor 2", "wor2"],
}

@dataclass
class TypeCurveMeta:
    key: tuple[str, str, str, str]
    lateral_display: str
    lateral_code: Any
    zone: str
    wps_display: str
    gsa_display: str
    wor: float
    address: str

@dataclass
class CaseDefinition:
    case_number: int
    case_id: str
    operator: str
    lateral_code: Any
    lateral_display: str
    lateral_ft: int
    wps: str
    zone: str
    gsa: str
    county: str
    type_curve: str
    wor: float
    ngl_yield: float
    gas_shrink: float
    drilling_cost: float
    facilities_cost: float
    flowlines_cost: float
    completion_cost: float
    drilling_month: float
    facilities_month: float
    flowlines_month: float
    completion_month: float
    pop_month: float
    extras: dict[str, Any]

@dataclass
class RunResult:
    case_number: int
    case_id: str
    operator: str
    lateral_length: str
    wps: str
    zone: str
    gsa: str
    county: str
    npv: Any
    aarr: Any
    cos: Any
    status: str
    error: str
    worker_id: int
    elapsed_seconds: float

def norm_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower().replace("&", " and ")
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"[^a-z0-9./$% ]+", "", s)
    return re.sub(r"\s+", " ", s).strip()

def norm_header(value: Any) -> str:
    s = norm_text(value)
    for ch in [",", ".", "$", "%"]:
        s = s.replace(ch, "")
    s = s.replace("/", " ")
    return re.sub(r"\s+", " ", s).strip()

def canonical_numeric_code(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    s = str(value).strip()
    try:
        f = float(s)
        if math.isfinite(f):
            if abs(f - round(f)) < 1e-10:
                return str(int(round(f)))
            return ("%.10f" % f).rstrip("0").rstrip(".")
    except Exception:
        pass
    return s.upper()

def norm_zone(value: Any) -> str:
    return str(value).strip().upper() if value is not None else ""

def tc_key(lateral: Any, zone: Any, wps: Any, gsa: Any) -> tuple[str, str, str, str]:
    return (
        canonical_numeric_code(lateral),
        norm_zone(zone),
        canonical_numeric_code(wps),
        canonical_numeric_code(gsa),
    )

def safe_case_token(value: Any) -> str:
    s = re.sub(r"\s+", "", str(value).strip())
    return s.replace("/", "-").replace("\\", "-").replace("|", "-")

def as_number(value: Any, field: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Missing numeric value for {field}")
    return float(value)

def display_cell(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    fmt = str(cell.number_format or "")
    if isinstance(v, (int, float)):
        if re.fullmatch(r"0+(?:\.0+)?", fmt):
            if "." in fmt:
                left, right = fmt.split(".", 1)
                return f"{float(v):0{len(left)+1+len(right)}.{len(right)}f}"
            return f"{int(round(float(v))):0{len(fmt)}d}"
        if float(v).is_integer():
            return str(int(v))
    return str(v).strip()

def lateral_to_feet(value: Any) -> int:
    f = float(value)
    if f >= 1000:
        return int(round(f))
    code = int(round(f))
    if code in LATERAL_FEET_MAP:
        return LATERAL_FEET_MAP[code]
    raise ValueError(f"Unknown lateral code {value!r}. Extend LATERAL_FEET_MAP if needed.")

def find_sheet(wb, wanted: str):
    target = norm_header(wanted)
    for ws in wb.worksheets:
        if norm_header(ws.title) == target:
            return ws
    for ws in wb.worksheets:
        if target in norm_header(ws.title) or norm_header(ws.title) in target:
            return ws
    raise KeyError(f"Sheet '{wanted}' not found.")

def find_header_row(ws, aliases: dict[str, list[str]], required_keys: Iterable[str], max_rows: int = 15):
    req = set(required_keys)
    for r in range(1, min(max_rows, ws.max_row) + 1):
        raw = {c: norm_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        found = {}
        for key, alias_list in aliases.items():
            alias_norms = [norm_header(a) for a in alias_list]
            for c, h in raw.items():
                if h and h in alias_norms:
                    found[key] = c
                    break
        if req.issubset(found):
            return r, found
    raise ValueError(f"Could not find required headers {sorted(req)} on '{ws.title}'.")

def nonblank_column_values(ws, header_row: int, col: int):
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(r, col)
        if cell.value in (None, ""):
            continue
        out.append((cell.value, display_cell(cell)))
    return out

def parse_matrix(ws):
    row, cols = find_header_row(
        ws, HEADER_ALIASES,
        ["operators", "lateral_lengths", "wps", "zone", "gsa", "county"]
    )
    return {k: nonblank_column_values(ws, row, cols[k])
            for k in ["operators", "lateral_lengths", "wps", "zone", "gsa", "county"]}

def parse_operator_assumptions(ws):
    row, cols = find_header_row(
        ws, HEADER_ALIASES,
        ["operators", "yield", "shrink", "drill_month", "fac_month",
         "flow_month", "comp_month", "pop_month"]
    )
    original_headers = {
        c: str(ws.cell(row, c).value).strip()
        for c in range(1, ws.max_column + 1)
        if ws.cell(row, c).value not in (None, "")
    }
    known_cols = set(cols.values())
    records = {}
    for r in range(row + 1, ws.max_row + 1):
        op = ws.cell(r, cols["operators"]).value
        if op in (None, ""):
            continue
        op_key = str(op).strip().upper()
        rec = {
            "yield": as_number(ws.cell(r, cols["yield"]).value, f"{op_key} yield"),
            "shrink": as_number(ws.cell(r, cols["shrink"]).value, f"{op_key} shrink"),
            "drill_month": as_number(ws.cell(r, cols["drill_month"]).value, f"{op_key} drilling timing"),
            "fac_month": as_number(ws.cell(r, cols["fac_month"]).value, f"{op_key} facilities timing"),
            "flow_month": as_number(ws.cell(r, cols["flow_month"]).value, f"{op_key} flowlines timing"),
            "comp_month": as_number(ws.cell(r, cols["comp_month"]).value, f"{op_key} completion timing"),
            "pop_month": as_number(ws.cell(r, cols["pop_month"]).value, f"{op_key} POP timing"),
            "_extras": {},
        }
        for c, header in original_headers.items():
            if c not in known_cols:
                val = ws.cell(r, c).value
                if val not in (None, ""):
                    rec["_extras"][header] = val
        records[op_key] = rec
    return records

def parse_gross_cost(ws):
    aliases = dict(HEADER_ALIASES)
    aliases["cost_lateral"] = ["lateral length", "lateral lengths", "ll"]
    row, cols = find_header_row(
        ws, aliases,
        ["operators", "cost_lateral", "drill_cost", "fac_cost", "flow_cost", "comp_cost"]
    )
    original_headers = {
        c: str(ws.cell(row, c).value).strip()
        for c in range(1, ws.max_column + 1)
        if ws.cell(row, c).value not in (None, "")
    }
    known_cols = set(cols.values())
    records = {}
    for r in range(row + 1, ws.max_row + 1):
        op = ws.cell(r, cols["operators"]).value
        ll = ws.cell(r, cols["cost_lateral"]).value
        if op in (None, "") or ll in (None, ""):
            continue
        rec = {
            "drill_cost": as_number(ws.cell(r, cols["drill_cost"]).value, "drilling cost"),
            "fac_cost": as_number(ws.cell(r, cols["fac_cost"]).value, "facilities cost"),
            "flow_cost": as_number(ws.cell(r, cols["flow_cost"]).value, "flowlines cost"),
            "comp_cost": as_number(ws.cell(r, cols["comp_cost"]).value, "completion cost"),
            "_extras": {},
        }
        for c, header in original_headers.items():
            if c not in known_cols:
                val = ws.cell(r, c).value
                if val not in (None, ""):
                    rec["_extras"][header] = val
        records[(str(op).strip().upper(), canonical_numeric_code(ll))] = rec
    return records

def parse_defaults(wb):
    try:
        ws = find_sheet(wb, "Defaults")
    except Exception:
        return {}
    result = {}
    for r in range(1, ws.max_row + 1):
        k, v = ws.cell(r, 1).value, ws.cell(r, 2).value
        if k in (None, "") or norm_header(k) in {"key", "name", "field"}:
            continue
        result[str(k).strip()] = v
    return result

def parse_type_curve_metadata(ws):
    row, cols = find_header_row(
        ws, HEADER_ALIASES,
        ["formation", "gsa", "wps", "tc_ll", "wor", "tc_address"],
        max_rows=20
    )
    result = {}
    for r in range(row + 1, ws.max_row + 1):
        formation = ws.cell(r, cols["formation"]).value
        ll_cell = ws.cell(r, cols["tc_ll"])
        wps_cell = ws.cell(r, cols["wps"])
        gsa_cell = ws.cell(r, cols["gsa"])
        address = ws.cell(r, cols["tc_address"]).value
        wor = ws.cell(r, cols["wor"]).value
        if formation in (None, "") or ll_cell.value in (None, "") or address in (None, ""):
            continue
        key = tc_key(ll_cell.value, formation, wps_cell.value, gsa_cell.value)
        meta = TypeCurveMeta(
            key=key,
            lateral_display=display_cell(ll_cell),
            lateral_code=ll_cell.value,
            zone=str(formation).strip(),
            wps_display=display_cell(wps_cell),
            gsa_display=display_cell(gsa_cell),
            wor=as_number(wor, f"WOR for {address}"),
            address=str(address).strip(),
        )
        if key in result:
            old = result[key]
            if old.address != meta.address or abs(old.wor - meta.wor) > 1e-9:
                raise ValueError(f"Conflicting type-curve metadata for {key}.")
        else:
            result[key] = meta
    if not result:
        raise ValueError("No type-curve metadata found.")
    return result

def extract_type_curve_table(ws):
    title_row = title_col = None
    for row in ws.iter_rows():
        for cell in row:
            if norm_header(cell.value) in ("type curves", "type curve"):
                title_row, title_col = cell.row, cell.column
                break
        if title_row:
            break

    # Fallback: no explicit "Type Curves" title cell — assume row 1 is the
    # header row directly (some sheets are laid out as a plain table with
    # no separate title banner above it).
    if not title_row:
        title_row, title_col = 0, 1

    header_row = title_row + 1
    start_col = title_col

    # A row is a plausible header row if it has several non-blank cells.
    # If the row right below the title is mostly blank (e.g. a spacer row,
    # merged banner, or units row), look a little further down instead of
    # assuming the layout and silently returning an empty table.
    def nonblank_count(r):
        return sum(1 for c in range(start_col, ws.max_column + 1) if ws.cell(r, c).value not in (None, ""))

    if nonblank_count(header_row) < 3:
        for candidate in range(header_row, min(header_row + 5, ws.max_row) + 1):
            if nonblank_count(candidate) >= 3:
                header_row = candidate
                break

    # Take every column from start_col through the last column that has a
    # header value OR any data beneath it, up to ws.max_column. No premature
    # cutoff on a short blank run — that heuristic is what caused this
    # table to come back empty when headers have any spacing/gaps.
    end_col = start_col
    for c in range(start_col, ws.max_column + 1):
        header_val = ws.cell(header_row, c).value
        has_data_below = any(
            ws.cell(r, c).value not in (None, "")
            for r in range(header_row + 1, min(header_row + 20, ws.max_row) + 1)
        )
        if header_val not in (None, "") or has_data_below:
            end_col = c
    headers = [ws.cell(header_row, c).value for c in range(start_col, end_col + 1)]

    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(start_col, end_col + 1)]
        if all(v in (None, "") for v in vals):
            continue
        data.append(vals)

    if not data:
        sample_rows = []
        for r in range(max(1, title_row - 1), min(title_row + 10, ws.max_row) + 1):
            sample_rows.append(
                f"  row {r}: " + repr([ws.cell(r, c).value for c in range(start_col, min(start_col + 8, ws.max_column) + 1)])
            )
        raise ValueError(
            "Paste-ready Type Curves table is empty.\n"
            f"Sheet: '{ws.title}' | title cell found at row {title_row}, col {title_col} "
            f"(0 means no 'Type Curves' title cell was found and row 1 was assumed).\n"
            f"Header row used: {header_row} | Headers read: {headers}\n"
            f"Columns scanned: {start_col} to {end_col}\n"
            "Nearby raw cell contents (for diagnosis):\n" + "\n".join(sample_rows)
        )
    return headers, data

def resolve_assumption(records, operator):
    op = operator.strip().upper()
    if op in records:
        return records[op]
    if "OTHER" in records:
        return records["OTHER"]
    raise KeyError(f"No assumptions for {operator} and no OTHER fallback.")

def resolve_cost(records, operator, lateral):
    op, ll = operator.strip().upper(), canonical_numeric_code(lateral)
    if (op, ll) in records:
        return records[(op, ll)]
    if ("OTHER", ll) in records:
        return records[("OTHER", ll)]
    raise KeyError(f"No gross cost for {operator}, lateral {lateral}, and no OTHER fallback.")

def build_cases(input_path: Path):
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    matrix = parse_matrix(find_sheet(wb, "Matrix"))
    op_assumptions = parse_operator_assumptions(find_sheet(wb, "ShrinkYield and Timing"))
    gross_cost = parse_gross_cost(find_sheet(wb, "Gross Cost"))
    tc_ws = find_sheet(wb, "Type Curves")
    tc_meta = parse_type_curve_metadata(tc_ws)
    defaults = parse_defaults(wb)

    operators = [(str(v).strip(), disp) for v, disp in matrix["operators"]]
    laterals = matrix["lateral_lengths"]
    wps_values = matrix["wps"]
    zones = [(str(v).strip(), disp) for v, disp in matrix["zone"]]
    gsas = matrix["gsa"]
    counties = [(str(v).strip(), disp) for v, disp in matrix["county"]]

    valid_tc = []
    missing = 0
    for ll_pair, zone_pair, wps_pair, gsa_pair in itertools.product(laterals, zones, wps_values, gsas):
        ll, ll_disp = ll_pair
        zone_raw, zone_disp = zone_pair
        wps_raw, wps_disp = wps_pair
        gsa_raw, gsa_disp = gsa_pair
        meta = tc_meta.get(tc_key(ll, zone_raw, wps_raw, gsa_raw))
        if meta is None:
            missing += 1
            continue
        valid_tc.append((ll, ll_disp, zone_disp, wps_disp, gsa_disp, meta))

    cases = []
    n = 1
    for combo, op_pair, county_pair in itertools.product(valid_tc, operators, counties):
        ll, ll_disp, zone_disp, wps_disp, gsa_disp, meta = combo
        operator, operator_disp = op_pair
        county, county_disp = county_pair
        op_rec = resolve_assumption(op_assumptions, operator)
        cost_rec = resolve_cost(gross_cost, operator, ll)
        extras = {}
        extras.update(op_rec.get("_extras", {}))
        extras.update(cost_rec.get("_extras", {}))
        extras.update(defaults)
        case_id = "_".join([
            safe_case_token(county_disp),
            safe_case_token(operator_disp),
            safe_case_token(ll_disp),
            safe_case_token(zone_disp),
            safe_case_token(wps_disp),
            safe_case_token(gsa_disp),
        ])
        cases.append(CaseDefinition(
            case_number=n,
            case_id=case_id,
            operator=operator_disp,
            lateral_code=ll,
            lateral_display=ll_disp,
            lateral_ft=lateral_to_feet(ll),
            wps=wps_disp,
            zone=zone_disp,
            gsa=gsa_disp,
            county=county_disp,
            type_curve=meta.address,
            wor=meta.wor,
            ngl_yield=float(op_rec["yield"]),
            gas_shrink=float(op_rec["shrink"]),
            drilling_cost=float(cost_rec["drill_cost"]),
            facilities_cost=float(cost_rec["fac_cost"]),
            flowlines_cost=float(cost_rec["flow_cost"]),
            completion_cost=float(cost_rec["comp_cost"]),
            drilling_month=float(op_rec["drill_month"]),
            facilities_month=float(op_rec["fac_month"]),
            flowlines_month=float(op_rec["flow_month"]),
            completion_month=float(op_rec["comp_month"]),
            pop_month=float(op_rec["pop_month"]),
            extras=extras,
        ))
        n += 1

    wb.close()
    if not cases:
        raise ValueError("No valid cases generated.")
    stats = {
        "operators": len(operators),
        "counties": len(counties),
        "type_curve_keys": len(tc_meta),
        "valid_tc_combinations": len(valid_tc),
        "missing_tc_combinations": missing,
        "generated_cases": len(cases),
    }
    return cases, stats

def import_excel_modules():
    import pythoncom
    import win32com.client
    return pythoncom, win32com.client

def wait_for_calculation(excel, pythoncom, timeout=600):
    start = time.monotonic()
    while int(excel.CalculationState) != XL_DONE:
        if time.monotonic() - start > timeout:
            raise TimeoutError(f"Excel calculation exceeded {timeout} sec.")
        pythoncom.PumpWaitingMessages()
        time.sleep(0.05)

def get_name_range(workbook, excel, name):
    try:
        return workbook.Names(name).RefersToRange
    except Exception:
        pass
    try:
        return excel.Range(name)
    except Exception as exc:
        available = ", ".join(sorted(n.Name for n in workbook.Names)) or "(none)"
        raise KeyError(
            f"SWELL workbook has no Named Range '{name}'. "
            f"Available Named Ranges: {available}. "
            f"Fix this in SWELL_CONFIG at the top of the script, or run --diagnose."
        ) from exc

def find_com_sheet(workbook, wanted):
    target = norm_header(wanted)
    for ws in workbook.Worksheets:
        if norm_header(ws.Name) == target:
            return ws
    raise KeyError(f"SWELL sheet '{wanted}' not found.")

def com_table_headers(lo):
    return {
        norm_header(lo.ListColumns(i).Name): (i, str(lo.ListColumns(i).Name))
        for i in range(1, lo.ListColumns.Count + 1)
    }

def find_wellattributes_table(ws):
    for i in range(1, ws.ListObjects.Count + 1):
        lo = ws.ListObjects(i)
        if norm_header(lo.Name) == "wellattributes":
            return lo
    for i in range(1, ws.ListObjects.Count + 1):
        lo = ws.ListObjects(i)
        headers = set(com_table_headers(lo))
        if "case" in headers and "type curve" in headers:
            return lo
    raise RuntimeError("Could not identify the WellAttributes table.")

def find_header_match(target_headers, aliases):
    for a in [norm_header(x) for x in aliases]:
        if a in target_headers:
            return target_headers[a]
    for h, info in target_headers.items():
        for a in [norm_header(x) for x in aliases]:
            if a and h and (a in h or h in a):
                return info
    return None

def resize_listobject_rows(lo, row_count):
    header = lo.HeaderRowRange
    ws = lo.Parent
    lo.Resize(ws.Range(
        ws.Cells(header.Row, header.Column),
        ws.Cells(header.Row + max(1, row_count), header.Column + lo.ListColumns.Count - 1)
    ))

def bulk_write_listcolumn(lo, col_index, values):
    rng = lo.ListColumns(col_index).DataBodyRange
    rng.Value = tuple((v,) for v in values)

def populate_wellattributes(workbook, cases):
    ws = find_com_sheet(workbook, "WellAttributes")
    lo = find_wellattributes_table(ws)
    resize_listobject_rows(lo, len(cases))
    headers = com_table_headers(lo)

    vals = {
        "case": [c.case_number for c in cases],
        "api": [c.case_id for c in cases],
        "project_name": [c.case_id for c in cases],
        "well_count": [1] * len(cases),
        "type_curve": [c.type_curve for c in cases],
        "reservoir": [c.zone for c in cases],
        "operator": [c.operator for c in cases],
        "wi": [1.0] * len(cases),
        "capital_wi": [1.0] * len(cases),
        "nri": [0.75] * len(cases),
        "yield": [c.ngl_yield for c in cases],
        "shrink": [c.gas_shrink for c in cases],
        "mmbtu": [1.0] * len(cases),
        "lateral_ft": [c.lateral_ft for c in cases],
        "drill_cost": [c.drilling_cost for c in cases],
        "fac_cost": [c.facilities_cost for c in cases],
        "flow_cost": [c.flowlines_cost for c in cases],
        "comp_cost": [c.completion_cost for c in cases],
        "art_lift_cost": [0.0] * len(cases),
        "gross_invest": [c.drilling_cost + c.facilities_cost + c.flowlines_cost + c.completion_cost for c in cases],
        "spud_date": [datetime(2027, 1, 1)] * len(cases),
        "drill_month": [c.drilling_month for c in cases],
        "fac_month": [c.facilities_month for c in cases],
        "flow_month": [c.flowlines_month for c in cases],
        "comp_month": [c.completion_month for c in cases],
        "pop_month": [c.pop_month for c in cases],
        "art_lift_years": [1.0] * len(cases),
        "obligation": ["REQUIRED WELL"] * len(cases),
        "tc_adj": [1.0] * len(cases),
        "county": [c.county for c in cases],
        "wor1": [c.wor for c in cases],
        "wor1_length": [1.0] * len(cases),
        "wor2": [c.wor for c in cases],
    }

    written = set()
    for key, colvals in vals.items():
        match = find_header_match(headers, WELLATTR_ALIASES[key])
        if match:
            idx, _ = match
            bulk_write_listcolumn(lo, idx, colvals)
            written.add(idx)

    extra_names = set()
    for c in cases:
        extra_names.update(c.extras)
    for extra in extra_names:
        target = headers.get(norm_header(extra))
        if not target:
            continue
        idx, _ = target
        if idx in written:
            continue
        bulk_write_listcolumn(lo, idx, [c.extras.get(extra) for c in cases])

    critical = ["case", "type_curve", "operator", "nri", "yield", "shrink", "lateral_ft", "county"]
    missing = [k for k in critical if not find_header_match(headers, WELLATTR_ALIASES[k])]
    if missing:
        raise RuntimeError(f"Missing required WellAttributes columns: {missing}")

def find_typecurve_table(ws):
    for i in range(1, ws.ListObjects.Count + 1):
        lo = ws.ListObjects(i)
        if "typecurve" in norm_header(lo.Name).replace(" ", ""):
            return lo
    if ws.ListObjects.Count:
        return ws.ListObjects(1)
    return None

def populate_type_curves(workbook, source_headers, source_data):
    ws = find_com_sheet(workbook, "Type Curves")
    lo = find_typecurve_table(ws)
    if lo is None:
        raise RuntimeError("SWELL Type Curves sheet has no Excel table.")
    resize_listobject_rows(lo, len(source_data))
    target_headers = com_table_headers(lo)
    matched = 0
    for src_idx, header in enumerate(source_headers):
        h = norm_header(header)
        if not h or h not in target_headers:
            continue
        idx, _ = target_headers[h]
        bulk_write_listcolumn(lo, idx, [row[src_idx] if src_idx < len(row) else None for row in source_data])
        matched += 1
    if matched < 3:
        raise RuntimeError("Input Type Curves table headers do not match SWELL Type Curves headers.")

def run_macro(workbook, excel):
    workbook.Activate()
    try:
        find_com_sheet(workbook, SWELL_CONFIG["dashboard_sheet"]).Activate()
    except Exception:
        pass
    last = None
    for macro in SWELL_CONFIG["macro_candidates"]:
        try:
            return excel.Run(f"'{workbook.Name}'!{macro}")
        except Exception as exc:
            last = exc
    raise RuntimeError(
        f"Could not run any of {SWELL_CONFIG['macro_candidates']} in the SWELL workbook: {last}. "
        f"Run --diagnose to list the macros Excel can actually see, then update "
        f"SWELL_CONFIG['macro_candidates']."
    )

def worker_run(worker_id, swell_path, input_path, case_dicts, worker_root, visible=False):
    pythoncom, win32com_client = import_excel_modules()
    cases = [CaseDefinition(**d) for d in case_dicts]

    input_wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    tc_headers, tc_data = extract_type_curve_table(find_sheet(input_wb, "Type Curves"))
    input_wb.close()

    root = Path(worker_root)
    root.mkdir(parents=True, exist_ok=True)
    source = Path(swell_path)
    worker_book = root / f"SWELL_worker_{worker_id:02d}{source.suffix}"
    shutil.copy2(source, worker_book)

    excel = workbook = None
    results = []
    pythoncom.CoInitialize()
    try:
        excel = win32com_client.DispatchEx("Excel.Application")
        excel.Visible = visible
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False
        excel.ScreenUpdating = False
        excel.Calculation = XL_CALCULATION_MANUAL

        workbook = excel.Workbooks.Open(
            str(worker_book), UpdateLinks=0, ReadOnly=False,
            IgnoreReadOnlyRecommended=True, Notify=False, AddToMru=False
        )

        t_setup = time.monotonic()
        populate_type_curves(workbook, tc_headers, tc_data)
        populate_wellattributes(workbook, cases)
        workbook.Activate()
        # CalculateFullRebuild is only needed ONCE, right after structurally
        # changing the workbook (resizing tables / repopulating data), so
        # Excel rebuilds its dependency tree correctly. Doing this per-case
        # instead of Calculate() is what causes multi-hour runtimes.
        excel.CalculateFullRebuild()
        wait_for_calculation(excel, pythoncom, 900)
        print(f"[worker {worker_id}] setup + initial full rebuild: {time.monotonic() - t_setup:.1f}s", flush=True)

        case_rng = get_name_range(workbook, excel, SWELL_CONFIG["case_name"])
        npv_rng = get_name_range(workbook, excel, SWELL_CONFIG["npv_name"])
        aarr_rng = get_name_range(workbook, excel, SWELL_CONFIG["aarr_name"])
        cos_rng = get_name_range(workbook, excel, SWELL_CONFIG["cos_input_name"])
        price_rng = get_name_range(workbook, excel, SWELL_CONFIG["price_scenario_name"])
        econ_rng = get_name_range(workbook, excel, SWELL_CONFIG["econ_limit_name"])

        # These three inputs are constant across every case in the run, so
        # they only need to be written once, not on every iteration.
        workbook.Activate()
        price_rng.Value = SWELL_CONFIG["price_scenario_value"]
        econ_rng.Value = SWELL_CONFIG["econ_limit_value"]
        cos_rng.Value = SWELL_CONFIG["cos_input_value"]

        case_times = []
        for c in cases:
            t0 = time.monotonic()
            try:
                workbook.Activate()
                case_rng.Value = c.case_number
                # Calculate() (not CalculateFull/CalculateFullRebuild) only
                # recalculates cells whose precedents actually changed, which
                # is the entire point of Manual calculation mode. This is the
                # single biggest speed lever in the whole run.
                excel.Calculate()
                wait_for_calculation(excel, pythoncom, 600)

                npv = npv_rng.Value
                aarr = aarr_rng.Value

                run_macro(workbook, excel)
                cos = cos_rng.Value

                elapsed = time.monotonic() - t0
                case_times.append(elapsed)
                results.append(asdict(RunResult(
                    c.case_number, c.case_id, c.operator, c.lateral_display,
                    c.wps, c.zone, c.gsa, c.county, npv, aarr, cos,
                    "SUCCESS", "", worker_id, elapsed
                )))
                if len(results) % 25 == 0:
                    avg = sum(case_times[-25:]) / len(case_times[-25:])
                    print(f"[worker {worker_id}] {len(results)}/{len(cases)} done "
                          f"| last-25 avg {avg:.2f}s/case", flush=True)
            except Exception as exc:
                results.append(asdict(RunResult(
                    c.case_number, c.case_id, c.operator, c.lateral_display,
                    c.wps, c.zone, c.gsa, c.county, None, None, None,
                    "FAILED", f"{type(exc).__name__}: {exc}", worker_id,
                    time.monotonic() - t0
                )))

        workbook.Close(SaveChanges=False)
        workbook = None
        excel.Quit()
        excel = None
    except Exception as exc:
        existing = {r["case_number"] for r in results}
        err = f"WORKER SETUP FAILURE: {type(exc).__name__}: {exc}"
        for c in cases:
            if c.case_number not in existing:
                results.append(asdict(RunResult(
                    c.case_number, c.case_id, c.operator, c.lateral_display,
                    c.wps, c.zone, c.gsa, c.county, None, None, None,
                    "FAILED", err, worker_id, 0.0
                )))
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return results

def split_evenly(items, n):
    n = max(1, min(n, len(items)))
    chunks = [[] for _ in range(n)]
    for i, item in enumerate(items):
        chunks[i % n].append(item)
    return [x for x in chunks if x]

def default_worker_count(case_count):
    cpu = os.cpu_count() or 4
    return max(1, min(6, max(2, cpu // 2), case_count))

def select_file(title, filetypes):
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        p = filedialog.askopenfilename(title=title, filetypes=filetypes)
        root.destroy()
        return Path(p) if p else None
    except Exception:
        return None

def choose_paths(input_arg, swell_arg):
    input_path = Path(input_arg).expanduser().resolve() if input_arg else None
    swell_path = Path(swell_arg).expanduser().resolve() if swell_arg else None
    if input_path is None:
        input_path = select_file("Select basin input workbook", [("Excel", "*.xlsx *.xlsm"), ("All files", "*.*")])
    if input_path is None:
        input_path = Path(input("Input workbook path: ").strip().strip('"')).expanduser().resolve()
    if swell_path is None:
        swell_path = select_file("Select SWELL model workbook (use a copy)", [("Excel Macro Workbook", "*.xlsm"), ("All files", "*.*")])
    if swell_path is None:
        swell_path = Path(input("SWELL workbook path: ").strip().strip('"')).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    if not swell_path.exists():
        raise FileNotFoundError(swell_path)
    return input_path, swell_path

def write_output(output_path, results, stats, run_type):
    results = sorted(results, key=lambda r: int(r["case_number"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["Case ID", "Operator", "Lateral Length", "WPS", "Zone", "GSA", "County", "NPV", "AARR", "CoS"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in results:
        ws.append([r["case_id"], r["operator"], r["lateral_length"], r["wps"], r["zone"],
                   r["gsa"], r["county"], r["npv"], r["aarr"], r["cos"]])

    log = wb.create_sheet("Run_Log")
    log.append(["Case ID", "Status", "Error", "Worker ID", "Elapsed Seconds"])
    for c in log[1]:
        c.font = Font(bold=True)
    for r in results:
        log.append([r["case_id"], r["status"], r["error"], r["worker_id"], r["elapsed_seconds"]])

    val = wb.create_sheet("Validation")
    val.append(["Item", "Value"])
    val["A1"].font = val["B1"].font = Font(bold=True)
    val.append(["Run Type", run_type])
    for k, v in stats.items():
        val.append([k, v])
    val.append(["Successful Cases", sum(r["status"] == "SUCCESS" for r in results)])
    val.append(["Failed Cases", sum(r["status"] != "SUCCESS" for r in results)])

    for col, width in {"A":42,"B":14,"C":16,"D":12,"E":14,"F":12,"G":14,"H":16,"I":14,"J":14}.items():
        ws.column_dimensions[col].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

def run_all(input_path, swell_path, cases, stats, output_path, workers=None, run_type="FULL", visible_workers=False):
    worker_count = workers or default_worker_count(len(cases))
    worker_count = max(1, min(worker_count, len(cases)))
    print(f"\nCases: {len(cases):,} | Excel workers: {worker_count}\nOutput: {output_path}\n")
    chunks = split_evenly(cases, worker_count)
    worker_root = Path(tempfile.mkdtemp(prefix="SWELL_parallel_"))
    all_results = []
    start = time.monotonic()
    try:
        with ProcessPoolExecutor(max_workers=worker_count) as pool:
            futures = {}
            for worker_id, chunk in enumerate(chunks, 1):
                fut = pool.submit(
                    worker_run, worker_id, str(swell_path), str(input_path),
                    [asdict(c) for c in chunk], str(worker_root), visible_workers
                )
                futures[fut] = worker_id
            completed = 0
            for fut in as_completed(futures):
                worker_id = futures[fut]
                rows = fut.result()
                all_results.extend(rows)
                completed += len(rows)
                success = sum(r["status"] == "SUCCESS" for r in rows)
                print(f"Worker {worker_id}: {success}/{len(rows)} successful | overall {completed}/{len(cases)}")
        write_output(output_path, all_results, stats, run_type)
        elapsed = time.monotonic() - start
        print(f"\nComplete in {elapsed/60:.1f} min -> {output_path}")
    finally:
        shutil.rmtree(worker_root, ignore_errors=True)

def make_output_path(input_path, suffix):
    return input_path.parent / f"{input_path.stem}_{suffix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

def diagnose_swell_workbook(swell_path):
    """Opens the SWELL workbook once, visibly, and prints everything needed
    to confirm/correct SWELL_CONFIG before committing to a full run."""
    pythoncom, win32com_client = import_excel_modules()
    pythoncom.CoInitialize()
    excel = workbook = None
    try:
        excel = win32com_client.DispatchEx("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(swell_path), UpdateLinks=0, ReadOnly=True)

        print("\n=== SHEETS ===")
        for ws in workbook.Worksheets:
            print(f"  {ws.Name!r}")

        print("\n=== NAMED RANGES ===")
        for n in workbook.Names:
            try:
                print(f"  {n.Name!r} -> {n.RefersTo}")
            except Exception:
                print(f"  {n.Name!r} -> <unreadable>")
        print(f"  (Config currently expects: {[SWELL_CONFIG[k] for k in ['case_name','npv_name','aarr_name','cos_input_name','price_scenario_name','econ_limit_name']]})")

        print("\n=== TABLES (ListObjects) PER SHEET ===")
        for ws in workbook.Worksheets:
            if ws.ListObjects.Count:
                for i in range(1, ws.ListObjects.Count + 1):
                    lo = ws.ListObjects(i)
                    cols = [lo.ListColumns(j).Name for j in range(1, lo.ListColumns.Count + 1)]
                    print(f"  [{ws.Name}] table {lo.Name!r}: {cols}")

        print("\n=== VBA MACROS (requires 'Trust access to the VBA project object model') ===")
        try:
            for comp in workbook.VBProject.VBComponents:
                mod = comp.CodeModule
                if mod.CountOfLines == 0:
                    continue
                for line_no in range(1, mod.CountOfLines + 1):
                    line = mod.Lines(line_no, 1).strip()
                    if line.lower().startswith("sub ") or line.lower().startswith("public sub "):
                        name = line.split("(")[0].split()[-1]
                        print(f"  {comp.Name}.{name}")
        except Exception as exc:
            print(f"  Could not enumerate macros ({exc}). Enable Trust access to the "
                  f"VBA project object model in Excel Trust Center Settings > Macro Settings, "
                  f"or just tell me the macro name(s) directly.")

        print(f"\n  Config currently expects macro_candidates: {SWELL_CONFIG['macro_candidates']}")
        print("\nUpdate SWELL_CONFIG at the top of this script to match anything that differs above.")
        input("\nPress Enter to close the diagnostic Excel window...")
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

def main(sample_fraction=1.0, sample_seed=42, output_suffix="SWELL_RESULTS", default_workers_override=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--input")
    parser.add_argument("--swell")
    parser.add_argument("--workers", type=int)
    parser.add_argument("--visible-workers", action="store_true")
    parser.add_argument("--diagnose", action="store_true",
                         help="Open the SWELL workbook and print its sheets, named ranges, "
                              "tables, and macros, then exit without running any cases.")
    args = parser.parse_args()

    if args.diagnose:
        swell_path = Path(args.swell).expanduser().resolve() if args.swell else None
        if swell_path is None:
            swell_path = select_file("Select SWELL model workbook", [("Excel Macro Workbook", "*.xlsm"), ("All files", "*.*")])
        if swell_path is None:
            swell_path = Path(input("SWELL workbook path: ").strip().strip('"')).expanduser().resolve()
        if not swell_path.exists():
            raise FileNotFoundError(swell_path)
        diagnose_swell_workbook(swell_path)
        return

    input_path, swell_path = choose_paths(args.input, args.swell)
    print("Reading input workbook and generating valid cases...")
    cases, stats = build_cases(input_path)
    print(f"Valid TC combinations: {stats['valid_tc_combinations']:,}")
    print(f"Generated cases: {len(cases):,}")
    print(f"Matrix combinations with no TC: {stats['missing_tc_combinations']:,}")

    run_type = "FULL"
    if sample_fraction < 1:
        n = max(1, math.ceil(len(cases) * sample_fraction))
        rng = random.Random(sample_seed)
        idx = sorted(rng.sample(range(len(cases)), n))
        cases = [cases[i] for i in idx]
        stats = dict(stats)
        stats["sample_fraction"] = sample_fraction
        stats["sampled_cases"] = len(cases)
        run_type = f"TEST_{sample_fraction:.1%}"
        print(f"Selected {len(cases):,} cases for {sample_fraction:.1%} test.")

    run_all(
        input_path, swell_path, cases, stats,
        make_output_path(input_path, output_suffix),
        workers=args.workers or default_workers_override,
        run_type=run_type,
        visible_workers=args.visible_workers,
    )

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    main()
